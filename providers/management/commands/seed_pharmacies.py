from django.core.management.base import BaseCommand
from django.db import transaction

from providers.models import Pharmacy
from openpyxl import load_workbook
from tqdm import tqdm


class Command(BaseCommand):
    help = "Seed pharmacies from XLSX with progress bar"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default="providers/data/pharmacies.xlsx",
            help="Path to XLSX file"
        )

    @transaction.atomic
    def handle(self, *args, **options):

        path = options["path"]

        workbook = load_workbook(filename=path, read_only=True)
        sheet = workbook.active

        total_rows = sheet.max_row - 1  # minus header

        pharmacies = []
        seen = set()

        created = 0
        skipped = 0

        rows = sheet.iter_rows(min_row=2, values_only=True)

        for row in tqdm(rows, total=total_rows, desc="Seeding pharmacies"):

            name, address, lat, lng, phone_numbers, website = row

            if not name or not address:
                skipped += 1
                continue

            name = str(name).strip()
            address = str(address).strip()

            lat = float(lat)
            lng = float(lng)

            unique_key = (name, lat, lng)

            if unique_key in seen:
                skipped += 1
                continue

            seen.add(unique_key)

            pharmacies.append(
                Pharmacy(
                    name=name,
                    address=address,
                    latitude=lat,
                    longitude=lng,
                    website=str(website).strip() if website else ""
                )
            )

        Pharmacy.objects.bulk_create(
            pharmacies,
            batch_size=1000,
            ignore_conflicts=True
        )

        created = len(pharmacies)

        self.stdout.write(
            self.style.SUCCESS(
                f"\n✅ Done!\nCreated: {created}\nSkipped: {skipped}\n"
            )
        )
